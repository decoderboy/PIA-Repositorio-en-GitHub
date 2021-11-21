import requests
import openpyxl
from bs4 import BeautifulSoup

# Obtencion de los datos mediante peticion GET
URL = "https://realpython.github.io/fake-jobs"
page = requests.get(URL)
# Analizamos el contenido HTML con BeautifulSoup
soup = BeautifulSoup(page.content, "html.parser")

results = soup.find(id="ResultsContainer")
#Buscar todos los elementos que el class "card-content"
job_elements = results.find_all("div", class_="card-content")
# Buscar todos los elementos que el h2 contenga en su texto
# la palabra "python"
python_jobs = results.find_all(
    "h2", string=lambda text: "python" in text.lower()
    )
# Buscar cada elemento que tenga referencia de python_hobs
# Y almacenarlo en python_jobs_elements
python_jobs_elements = [
    h2_element.parent.parent.parent for h2_element in python_jobs
    ]
book = openpyxl.Workbook()
sheet = book.active
sheet.merge_cells('A1:C1')
sheet.merge_cells('D1:F1')
sheet.merge_cells('G1:J1')
sheet.merge_cells('K1:S1')
sheet['A1'] = "Company"
sheet['D1'] = "Location"
sheet['G1'] = "Title"
sheet['K1'] = "Apply Here"
i = 1
# Mostrar informacion de python_jobs_elements
for job_element in python_jobs_elements:
    i = i + 1
    sheet.merge_cells(f'A{i}:C{i}')
    sheet.merge_cells(f'D{i}:F{i}')
    sheet.merge_cells(f'G{i}:J{i}')
    sheet.merge_cells(f'K{i}:S{i}')
    title_element = job_element.find("h2", class_="title")
    company_element = job_element.find("h3", class_="company")
    location_element = job_element.find("p", class_="location")
    # De la lista de elementos de la etiqueta "a" buscamos
    # el primer elemento que incluya href.
    link_url = job_element.find_all("a")[1]["href"]
    print(company_element.text.strip())
    print(location_element.text.strip())
    print(title_element.text.strip())
    # Imprimimos la salida con link_url
    sheet["A"+str(i)] = str(company_element.text.strip())
    sheet["D"+str(i)] = str(location_element.text.strip())
    sheet["G"+str(i)] = str(title_element.text.strip())
    sheet["K"+str(i)] = str(link_url)
    print(f"Apply Here: {link_url}\n")
    print()

book.save('python_fake_jobs.csv')
'# Author: Ivan Eduardo Reyes Garcia'
'# Matricula: 1918865'
 
